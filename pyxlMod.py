#!/usr/bin/env python
# coding: utf-8


import os
import sys
import openpyxl
from openpyxl.cell.cell import get_column_letter, column_index_from_string


'''
openpyxl模块：
1.可以
'''


# 格式化路径。
def formatPath(path):
    return os.path.normpath(path)


############# decorator################
def columLetterToIndex(func):
    def translate(*args):
        if "getColValues" == func.__name__:
            colNum = args[1] if isinstance(args[1], int) else letterToInt(args[1]) 
            sheetIndex = None if len(args) == 2 else args[2]
            return func(args[0], colNum, sheetIndex)
        elif "getCellValue" == func.__name__:
            rowNum = args[1] if isinstance(args[1], int) else letterToInt(args[1])
            colNum = args[2] if isinstance(args[2], int) else letterToInt(args[2]) 
            sheetIndex = None if len(args) == 3 else args[3]
            return  func(args[0], rowNum, colNum, sheetIndex)
        else:
            return
    return translate


# 自己写的表格字母数字互转，后面发现模块自带。。。。。。。。
def intToLetter(index_int):
    letterAll = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if 0 >= index_int:
        print "column index can't less 0 or equal 0"
        sys.exit()
    elif 26 > index_int:
        return letterAll[index_int - 1]
    else:
        quotient = index_int / 26
        remainder = index_int % 26
        result = letterAll[remainder - 1]
        result = intToLetter(quotient) + result
    return result


def letterToInt(str_):
    letterAll = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    int_list = []
    int_ = 0
    for i in range(len(str_)):
        if str_[i] not in letterAll:
            print i + "not in " + letterAll
            sys.exit()
        else:
            index_ = letterAll.index(str_[i])
            int_ = int_ + (index_ + 1) * (26 ** (len(str_) - i - 1))
    return int_
############################################


class ReadExcel(object):

    def __init__(self, filePath):
        self.filePath = os.path.normpath(filePath)
        self.excel = openpyxl.load_workbook(filePath, read_only=True)
        self.activeSheetIndex = 0
        self.sheet = None
        self.setActiveSheet(0)
# 关闭
    def close(self):
        self.excel.close()


# 设置操作sheet页
    def setActiveSheet(self, sheetIndex):
        sheet_name = self.excel.get_sheet_names()[sheetIndex]
        self.activeSheetIndex = sheetIndex
        self.sheet = self.excel.get_sheet_by_name(sheet_name)

# 
    def __getSheetByIndex(self, sheetIndex = None):
        if sheetIndex == None or sheetIndex == self.activeSheetIndex:
            sheetObj = self.sheet
        else:
            sheet_name = self.excel.get_sheet_names()[sheetIndex]
            sheetObj = self.excel.get_sheet_by_name(sheet_name)
        return sheetObj

# 获取但前sheet页的名字
    def getSheetNames(self):
        return self.excel.get_sheet_names()

# 获取单元格数据，有装饰器把字母列标转化成索引
# 有简单方法，只是为了练习装饰器
    @columLetterToIndex
    def getCellValue(self, rowNum, colNum, sheetIndex = None):
        sheetObj = self.__getSheetByIndexsheetIndex()
        return sheetObj.cell(row = rowNum, column = colNum).value


# 获取整列数据，有装饰器同上
    @columLetterToIndex
    def getColValues(self,colNum, sheetIndex = None):
        count = 1
        sheetObj = self.__getSheetByIndex(sheetIndex)
        colObj = sheetObj[str(colNum)]
        result = []
        for j in colObj:
            result.append(j.value)
        return result

# 获取一行数据
    def getRowValues(self,rowNum, sheetIndex = None):
        sheetObj = self.__getSheetByIndex(sheetIndex)
        rowObj = sheetObj[str(rowNum)]
        result = []
        for j in rowObj:
            result.append(j.value)
        return result


# 获取一块区域的数据
    def getRangeValues(self, startCell, endCell, sheetIndex = None):
        data = []
        sheetObj = self.__getSheetByIndex(sheetIndex)
        rangeObj = sheetObj[startCell:endCell]
        for i in rangeObj:
            data.append([])
            for j in i:
                data[-1].append(j.value)
        return data


# 获取全部数据
    def getAllValues(self, sheetIndex = None):
        allData = []
        sheetObj = self.__getSheetByIndex(sheetIndex)
        rowsValue = sheetObj.rows
        for i in rowsValue:
            allData.append([])
            for j in i:
                allData[-1].append(j.value)
        return allData


# 获取行数
    def getRowCounts(self, sheetIndex = None):
        sheetObj = self.__getSheetByIndex(sheetIndex)
        count = 0
        for i in sheetObj.rows:
            count += 1
        return count

# 获取列数
    def getColumnCounts(self, sheetIndex = None):
        sheetObj = self.__getSheetByIndex(sheetIndex)
        count = 0
        for i in sheetObj.columns:
            count += 1
        return count


class WriteExcel(object):

    def __init__(self, excelPath):
        self.filePath = excelPath
        self.wrObj = self.creat_or_open()
        self.sheetNames = self.__getSheetIndex_Names()


    def __getSheetIndex_Names(self):
        names = self.wrObj.get_sheet_names()
        return {x:names[x] for x in range(len(names))}


# 新建或加载一个表格，存在则加载，不存在则新建。
    def creat_or_open(self):
        if os.path.isfile(self.filePath):
            return openpyxl.load_workbook(self.filePath)
        else:
            return openpyxl.Workbook()

# 创建sheet页
    def creatSheet(self, title='new sheet', index__=None):
        if title in self.sheetNames.values():
            raise Exception('title name: \"' + title + '\" is exist!')
        index__ = len(self.sheetNames) if index__ is None else index__
        sheet = self.wrObj.create_sheet(index=index__, title=title)
        self.sheetNames = self.__getSheetIndex_Names()
        return sheet

# 获取sheet对象，参数为sheet页索引
    def getSheet_by_index(self, index__=0):
        sheet = self.wrObj.get_sheet_by_name(self.sheetNames[index__])
        return sheet

# 插入一行数据
    def appendRow(self, rowValue, sheetObj):
        sheetObj.append(rowValue)

# 更新单元格数据
    def updateCell(self, cell_value_tuple, sheetObj):
        if type(cell_value_tuple) is tuple:
            sheetObj[cell_value_tuple[0]] = cell_value_tuple[1]
        else:
            print "parament 1 is a tuple like ('A1','value')"

# 删除sheet 页
    def delectSheet_by_index(self, index__):
        del self.wrObj[self.sheetNames[index__]]
        self.sheetNames = self.__getSheetIndex_Names()

# 设置行高（测试没有效果，原因不明）
    def setRowHeight(self, sheetObj, rowNum_Int, height=None):
        if heigth is not None:
            sheetObj.row_dimensions[rowNum].heigth = height
        else:
            pass

# 设置列宽（测试没有效果，原因不明）
    def setColWidth(self, sheetObj, col_Str, width=None):
        if width is not None:
            sheetObj.column_dimensions[col_Str].width = width
        else:
            pass

# 保存表格，如果传路路径则保存为该路径，否则保存为实例化该对象时传入的路径
    def saveExcel(self, filePath=None):
        if filePath is None:
            self.wrObj.save(self.filePath)
        else:
            self.wrObj.save(filePath)



if __name__ == "__main__":
    wObj = WriteExcel('test.xlsx')
    # wObj.creatSheet(title='new sheet22')
    print wObj.sheetNames
    # sheet = wObj.getSheet_by_index(0)
    # wObj.delectSheet_by_index(0)
    wObj.saveExcel()

    wObj.close()


